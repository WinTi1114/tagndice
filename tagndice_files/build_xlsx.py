# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ---------- palette (matches the PDF: parchment / bordeaux / gold) ----------
BORDEAUX = "7A1F2B"
BORDEAUX_D = "5C1620"
GOLD = "8A6A2F"
INPUT_FILL = "FFF7CC"
GRAY_CALC = "E7E2D6"
GM_FILL = "EFE3EC"
WHITE = "FFFFFF"
FONT_BASE = "Arial"


def f(size=10, bold=False, italic=False, color="241A12"):
    return Font(name=FONT_BASE, size=size, bold=bold, italic=italic, color=color)


thin = Side(style="thin", color="B7A37E")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def title_band(ws, row, col1, col2, text, height=26):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.font = f(13, bold=True, color=WHITE)
    c.fill = fill(BORDEAUX)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    for cc in range(col1, col2 + 1):
        ws.cell(row=row, column=cc).fill = fill(BORDEAUX)


def sec_band(ws, row, col1, col2, text, height=16):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value="◆ " + text)
    c.font = f(10.5, bold=True, color=WHITE)
    c.fill = fill(GOLD)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    for cc in range(col1, col2 + 1):
        ws.cell(row=row, column=cc).fill = fill(GOLD)


def label_cell(ws, row, col, text, bold=True, align="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = f(9, bold=bold, color=BORDEAUX_D)
    c.alignment = Alignment(horizontal=align, vertical="center")
    return c


def input_cell(ws, row, col, value=None, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(INPUT_FILL)
    c.border = border_all
    c.font = f(10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    return c


def calc_cell(ws, row, col, formula=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = fill(GRAY_CALC)
    c.border = border_all
    c.font = f(10, bold=True, color=BORDEAUX_D)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c


def note(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.font = f(8.5, italic=True, color="6B5B45")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return c


wb = Workbook()

# =========================================================================
# SHEET 1 : 앞면 (액션 시트)
# =========================================================================
ws = wb.active
ws.title = "앞면(액션시트)"
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGH", [13, 20, 20, 13, 13, 20, 13, 15]):
    ws.column_dimensions[col].width = w

r = 1
title_band(ws, r, 1, 8, "TAG & DICE — 캐릭터 액션 시트"); r += 1
note(ws, r, 1, 8, "안내 : 노란색 칸에 내용을 입력하세요. 회색 칸(레벨 / 등급 합)은 수식으로 자동 계산되니 직접 입력하지 마세요.")
r += 2

label_cell(ws, r, 1, "PC 이름")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
input_cell(ws, r, 2)
label_cell(ws, r, 6, "레벨(자동)")
lvl_row = r
calc_cell(ws, r, 7)
r += 1
label_cell(ws, r, 1, "종족"); input_cell(ws, r, 2)
label_cell(ws, r, 3, "배경"); input_cell(ws, r, 4)
label_cell(ws, r, 5, "본능"); input_cell(ws, r, 6)
label_cell(ws, r, 7, "경험태그 등급합(자동)")
sum_row = r
calc_cell(ws, r, 8)
r += 2

sec_band(ws, r, 1, 8, "태생 태그 — 종족·배경·본능 (고정 1등급, 모든 판정에 상시 적용 · 레벨 합산에는 미포함)"); r += 1
label_cell(ws, r, 1, "구분")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
label_cell(ws, r, 2, "태그 이름")
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
label_cell(ws, r, 3, "내용")
r += 1
for kind in ["종족", "배경", "본능"]:
    label_cell(ws, r, 1, kind, bold=False)
    input_cell(ws, r, 2)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    input_cell(ws, r, 3)
    r += 1
r += 1

sec_band(ws, r, 1, 8, "경험 태그 — 지식·기능 / 캐릭터 생성 시 총 3포인트 배분 / 캠페인 중 최대 10개까지 확장"); r += 1
label_cell(ws, r, 1, "유형")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
label_cell(ws, r, 2, "태그 이름")
label_cell(ws, r, 7, "등급(1~3)")
r += 1
note(ws, r, 1, 8, "예시 (입력칸 아님) :  지식 | 고대 문자 해독 | 1   ·   기능 | 절벽 등반 | 1")
r += 1

dv_type = DataValidation(type="list", formula1='"지식,기능"', allow_blank=True)
dv_grade = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
ws.add_data_validation(dv_type)
ws.add_data_validation(dv_grade)

exp_first = r
for i in range(10):
    input_cell(ws, r, 1)
    dv_type.add(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    input_cell(ws, r, 2)
    input_cell(ws, r, 7)
    dv_grade.add(ws.cell(row=r, column=7))
    r += 1
exp_last = r - 1
r += 1

# level / grade-sum formulas (now that the experience-tag grade range is known)
ws.cell(row=sum_row, column=8, value=f"=SUM(G{exp_first}:G{exp_last})")
ws.cell(row=sum_row, column=8).comment = Comment(
    f"경험 태그 등급 열(G열, G{exp_first}:G{exp_last})의 합계입니다. "
    "태생 태그(종족·배경·본능)는 상시 적용되는 고정 효과일 뿐, 이 등급 합에는 포함되지 않습니다.", "spec")
ws.cell(row=lvl_row, column=7, value=(
    f'=IF(H{sum_row}>=21,"Ⅴ",IF(H{sum_row}>=16,"Ⅳ",'
    f'IF(H{sum_row}>=11,"Ⅲ",IF(H{sum_row}>=6,"Ⅱ","Ⅰ"))))'
))
ws.cell(row=lvl_row, column=7).comment = Comment(
    "경험 태그 등급 합계 기준 : Ⅰ 1~5 · Ⅱ 6~10 · Ⅲ 11~15 · Ⅳ 16~20 · Ⅴ 21+", "spec")

sec_band(ws, r, 1, 4, "핵심 자원"); sec_band(ws, r, 5, 8, "판정 & 태그 선언"); r += 1
res_top = r
cond_note = {"정신력": "지적 행동에 사용", "피로도": "신체적 행동에 사용"}
for name, start in [("체력", 5), ("정신력", 3), ("피로도", 3)]:
    label_cell(ws, r, 1, name)
    label_cell(ws, r, 2, "현재", bold=False, align="right")
    input_cell(ws, r, 3, start, wrap=False)
    if name in cond_note:
        note(ws, r, 4, 4, cond_note[name])
    r += 1
label_cell(ws, r, 1, "(참고)", bold=False)
note(ws, r, 2, 4, "최대치는 레벨업 시 +1 (체력/정신력/피로도 중 택1)")
r += 1
label_cell(ws, r, 1, "(상처)", bold=False)
note(ws, r, 2, 4, "피해 시 체력 칸에 상처 표시(페널티 없음) · 전부 채워지면 즉시 행동불능")
res_bottom = r
r += 1

r2 = res_top
note(ws, r2, 5, 8, "판정 : 2d6 → 10+ 거의 문제없이 성공 · 5~9 성공하나 침식 발동/GM 지정 위험 · 4- GM 재량(대개 실패)"); r2 += 1
note(ws, r2, 5, 8, "태그 선언 : 관련 컨디션(정신력/피로도) 1점 소모 → 판정 없이 자동 성공."); r2 += 1
note(ws, r2, 5, 8, "컨디션이 0이면 태그 선언 불가 → 2d6 판정으로 대신합니다."); r2 += 1
note(ws, r2, 5, 8, "태그 등급 : 1급 견습 · 2급 숙련 · 3급 달인"); r2 += 1
r = max(res_bottom, r2) + 2

sec_band(ws, r, 1, 8, "침식 태그 — 사경 판정(2d6) 5~9 결과로 획득, 최대 3개 (레벨업 시 1개 정화 가능)"); r += 1
for _ in range(3):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    input_cell(ws, r, 1)
    r += 1
r += 1

sec_band(ws, r, 1, 8, "마법 물품 — 나의 정체성을 상징하는 특별한 물품. 세션당 3회, 내 차례에 1회만 발동"); r += 1
label_cell(ws, r, 1, "물품 이름")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
input_cell(ws, r, 2)
label_cell(ws, r, 5, "형태")
dv_miform = DataValidation(type="list", formula1='"무기류,방어구류,마법도구류"', allow_blank=True)
ws.add_data_validation(dv_miform)
input_cell(ws, r, 6, wrap=False)
dv_miform.add(ws.cell(row=r, column=6))
label_cell(ws, r, 7, "사용(세션당)")
dv_miuse = DataValidation(type="list", formula1='"☐☐☐,☑☐☐,☑☑☐,☑☑☑"', allow_blank=True)
ws.add_data_validation(dv_miuse)
input_cell(ws, r, 8, "☐☐☐", wrap=False)
dv_miuse.add(ws.cell(row=r, column=8))
r += 1
label_cell(ws, r, 1, "전용 태그")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
input_cell(ws, r, 2)
r += 1
note(ws, r, 1, 8, "이 물품의 태그는 캐릭터의 다른 태그와 유일하게 중첩 가능 — 경험 태그로 판정/선언할 때 함께 걸어 두 효과를 하나의 묘사로 엮을 수 있습니다.")
r += 2

sec_band(ws, r, 1, 5, "착용 장비 & 소지품 태그 — 착용 장비·마법 물품은 소지품 한도 제외")
sec_band(ws, r, 6, 8, "화폐")
r += 1
label_cell(ws, r, 1, "무기")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
input_cell(ws, r, 2, wrap=False)
label_cell(ws, r, 4, "방어구")
input_cell(ws, r, 5, wrap=False)
fx_label = label_cell(ws, r, 6, "금화"); input_cell(ws, r, 7, 0, wrap=False)
fx_label.comment = Comment(
    "환율 1금=10은=100동 · 기준: 평민 노동자 하루 일당=1금. "
    "동/은/금은 일상 경제 전반에, 조각은 몬스터·마법이 얽힌 거래에만 사용합니다.", "spec")
r += 1
note(ws, r, 1, 5, "소지품 태그 — 경험 태그와 별개 슬롯, 최대 10개. 착용 장비·마법 물품·사소한 소지품(약간의 여비·기본 야영 도구 등)은 한도 제외. 자리가 다 찼으면 기존 태그 하나를 버리고(즉시 가능) 새로 획득.")
label_cell(ws, r, 6, "은화"); input_cell(ws, r, 7, 0, wrap=False)
r += 1
label_cell(ws, r, 1, "#")
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
label_cell(ws, r, 2, "소지품 태그")
label_cell(ws, r, 6, "동화"); input_cell(ws, r, 7, 0, wrap=False)
r += 1
gear_start = r
for i in range(10):
    idx = ws.cell(row=r, column=1, value=i + 1)
    idx.font = f(9, color="6B5B45")
    idx.alignment = Alignment(horizontal="center", vertical="center")
    idx.border = border_all
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    input_cell(ws, r, 2)
    if i == 0:
        label_cell(ws, r, 6, "조각"); input_cell(ws, r, 7, 0, wrap=False)
    r += 1

ws.freeze_panes = "A4"

# =========================================================================
# SHEET 2 : 뒷면 (백스토리 카드)
# =========================================================================
ws2 = wb.create_sheet("뒷면(백스토리카드)")
ws2.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGH", [13, 20, 20, 13, 13, 20, 13, 15]):
    ws2.column_dimensions[col].width = w

r = 1
title_band(ws2, r, 1, 8, "TAG & DICE — 캐릭터 백스토리 카드"); r += 1
label_cell(ws2, r, 1, "PC 이름")
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
input_cell(ws2, r, 2)
r += 2

sec_band(ws2, r, 1, 8, "기본 정보"); r += 1
label_cell(ws2, r, 1, "가치관"); ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); input_cell(ws2, r, 2)
label_cell(ws2, r, 4, "나이"); input_cell(ws2, r, 5)
label_cell(ws2, r, 6, "키/몸무게"); ws2.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8); input_cell(ws2, r, 7)
r += 1
label_cell(ws2, r, 1, "외모 특징"); ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8); input_cell(ws2, r, 2)
r += 1
label_cell(ws2, r, 1, "초상화 링크"); ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8); input_cell(ws2, r, 2)
note(ws2, r, 2, 2, "")  # keep formatting simple; cell already has input fill
r += 2

sec_band(ws2, r, 1, 8, "상세 백스토리"); r += 1
story_top = r
ws2.merge_cells(start_row=r, start_column=1, end_row=r + 13, end_column=8)
sc = ws2.cell(row=r, column=1, value="")
sc.fill = fill(INPUT_FILL)
sc.border = border_all
sc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
for rr in range(r, r + 14):
    ws2.row_dimensions[rr].height = 18
r += 15

sec_band(ws2, r, 1, 5, "마스터 비밀 메모 (GM 전용 — 플레이어는 참고하지 마세요)")
sec_band(ws2, r, 6, 8, "")
r += 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r + 5, end_column=5)
gm = ws2.cell(row=r, column=1, value="")
gm.fill = fill(GM_FILL)
gm.border = border_all
gm.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
gm_top = r
ws2.merge_cells(start_row=r, start_column=6, end_row=r + 5, end_column=8)
sess_note = r
for rr in range(r, r + 6):
    ws2.row_dimensions[rr].height = 18
r += 7

sec_band(ws2, r, 1, 8, "규칙 요약본"); r += 1
cheat_lines = [
    "판정 2d6 — 10+ 거의 문제없이 성공 · 5~9 성공하나 침식 발동/GM 지정 위험 · 4- GM 재량(대개 실패)",
    "태그 선언 — 컨디션(정신력/피로도) 1점 소모 → 판정 없이 자동 성공. 0이면 불가.",
    "컨디션 — 정신력 : 지적 행동에 사용 · 피로도 : 신체적 행동에 사용. 같은 태그도 행동 묘사에 따라 어떤 컨디션을 쓸지는 마스터가 판단.",
    "태그 등급 — 1급 견습 · 2급 숙련 · 3급 달인 (등급이 오를수록 수식어를 이어붙여 이름 확장)",
    "레벨 — 경험 태그 등급 합 : Ⅰ 1~5 · Ⅱ 6~10 · Ⅲ 11~15 · Ⅳ 16~20 · Ⅴ 21+ (레벨업 시 체력/정신력/피로도 중 1개 최대치 +1, 침식 태그 1개 정화 가능)",
    "상처 — 피해를 받으면 체력 칸에 상처 표시(페널티 없음). 모든 칸이 상처로 가득 차면 즉시 행동불능.",
    "사경 판정(2d6) — 10+ 생존 · 5~9 생존하되 침식 태그 획득(최대 3개) · 4- 사망. 침식 태그 3개 보유 중 다시 행동불능이 되면 굴림 없이 사망.",
    "야영 — 제대로 자리 잡고 쉴 때 일행 중 1인이 보급품/식량 1회분 소비. 짧은 휴식 : 정신력·피로도 중 택1 전부 회복 · 긴 휴식 : 정신력·피로도 모두 전부 회복",
    "치료 — 상처 치료에 도움이 되는 경험 태그를 선언하면 체력 1칸 회복(상처 1개 제거)",
    "화폐 — 동/은/금 : 일상 경제 전반 · 조각 : 몬스터·마법 관련 거래 전용 · 환율 1금=10은=100동 (기준: 평민 하루 일당=1금)",
]
# two side-by-side columns (left: cols 1-4, right: cols 5-8) so the summary
# uses the full row width instead of stacking every line in one narrow column
# — replaces the old single-column stack + the now-removed 세션 기록 table.
half = (len(cheat_lines) + 1) // 2
for i in range(half):
    note(ws2, r, 1, 4, "• " + cheat_lines[i])
    right_idx = i + half
    if right_idx < len(cheat_lines):
        note(ws2, r, 5, 8, "• " + cheat_lines[right_idx])
    r += 1
r += 1

ws2.freeze_panes = "A3"

for sheet in (ws, ws2):
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("/home/claude/tagndice/tagndice_character_sheet.xlsx")
print("workbook built OK, sheet1 rows used:", exp_last, " final row sheet2:", r)
